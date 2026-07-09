import asyncio
import io
import openpyxl
from Backend.app.services.teams_client import graph
from Backend.app.routers.excel import _encode_share_url

async def main():
    url = 'https://usilparaguay-my.sharepoint.com/:x:/g/personal/resteche_usil_edu_py/IQCHMuoLYGs9T4NDeid5n9A7AZvphg9oml_g9dt-GYD5tY0?e=zNyWPr'
    encoded_url = _encode_share_url(url)
    try:
        contents = await graph.get_raw(f"/shares/{encoded_url}/driveItem/content")
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        print("Sheets found:", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\\n--- {sheet_name} ---")
            
            # Find header row (same logic as our preview functions)
            header_row_idx = None
            headers = {}
            for row_idx in range(1, min(10, ws.max_row + 1)):
                row_vals = [str(ws.cell(row=row_idx, column=c).value or "").strip() for c in range(1, min(20, ws.max_column + 1))]
                valid_count = sum(1 for v in row_vals if v)
                if valid_count >= 3:
                    header_row_idx = row_idx
                    headers = {v: c for c, v in enumerate(row_vals, start=1) if v}
                    break
            
            if header_row_idx:
                print(f"Header Row: {header_row_idx}")
                print(f"Headers: {list(headers.keys())}")
            else:
                print("No headers found in first 10 rows.")
                
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    asyncio.run(main())
