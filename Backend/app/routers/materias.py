"""Matriculación masiva por materias, desde una planilla de OneDrive con
una hoja por alumno (columnas Curso / Día / Hora + "Nombre y Apellido:" /
"Cédula:" en celdas sueltas — formato real de uso, no una tabla plana).

El problema que resuelve: el nombre de la materia tipeado a mano en la
planilla casi nunca coincide letra por letra con el nombre real del curso
en Canvas (mayúsculas, tildes, "II" vs "2", abreviaturas). Antes de esto,
el flujo de matriculación por nombre exigía coincidencia EXACTA y, si no
la encontraba, terminaba creando un curso nuevo duplicado en vez de
matricular en el existente (ver canvas_client.search_course_by_name).

Acá en cambio se hace fuzzy matching contra la lista real de cursos de la
cuenta, y SIEMPRE se muestra una vista previa con el score de cada match
antes de matricular a nadie — nunca se crea un curso nuevo desde este
flujo, si no hay un match razonable se marca para revisión manual.
"""
import difflib
import asyncio
import json
import logging
import re
import unicodedata
from typing import Any

import openpyxl
from fastapi import APIRouter, BackgroundTasks, HTTPException
from pydantic import BaseModel

from app.core.config import settings
from app.core import jobs
from app.services import canvas_client as canvas
from app.services import teams_client as graph
from app.routers.excel import UrlOnlyRequest, _encode_share_url

router = APIRouter(prefix="/materias", tags=["Matriculación por Materias"])
logger = logging.getLogger(__name__)
_ACCOUNT = settings.canvas_account_id

# Por debajo de este score el match se marca "sin_match" y no se sugiere nada.
REVIEW_THRESHOLD = 0.55
# Por encima de este score el match se marca "auto" (alta confianza).
AUTO_THRESHOLD = 0.85

_NOMBRE_LABELS = ["nombre y apellido", "nombre completo", "alumno"]
_CEDULA_LABELS = ["cedula", "ci", "documento"]
_CURSO_HEADER = "curso"


def _normalize(s: str) -> str:
    """Normaliza texto para comparar sin importar tildes/mayúsculas/espacios
    ni puntuación suelta (ej. 'Investigación y Análisis de Mercado II' ==
    'investigacion y analisis de mercado 2' en similitud, no en igualdad)."""
    ascii_s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii")
    ascii_s = re.sub(r"[^a-z0-9\s]", " ", ascii_s.lower())
    return " ".join(ascii_s.split())


def _score(a: str, b: str) -> float:
    """Similitud entre dos nombres de curso. Además del ratio de
    SequenceMatcher, si uno es substring completo del otro A NIVEL DE
    PALABRA (típico cuando el nombre real en Canvas trae un sufijo de
    código/período, ej. 'Estrategia de Producto' vs 'Estrategia de
    Producto - CPEL 2026') se sube el piso a 0.9 — un ratio puro penaliza
    injustamente ese caso por la diferencia de longitud.

    La contención se exige a nivel de PALABRA completa, no de caracteres
    sueltos: comparar substrings de caracteres da falsos positivos graves
    con nombres cortos (ej. 'Etica' contiene literalmente 'tic', pero no
    tiene nada que ver con la materia 'TIC'). Además, para que cuente una
    sola palabra de contención, esa palabra tiene que ser larga (6+
    caracteres) — si no, cualquier palabra corta y genérica calificaría."""
    na, nb = _normalize(a), _normalize(b)
    if not na or not nb:
        return 0.0
    base = difflib.SequenceMatcher(None, na, nb).ratio()
    wa, wb = na.split(), nb.split()
    shorter, longer = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    shorter_str, longer_str = " ".join(shorter), " ".join(longer)
    if shorter_str and shorter_str in longer_str and (len(shorter) >= 2 or len(shorter_str) >= 6):
        base = max(base, 0.9)
    return base


def _best_matches(materia_name: str, courses: list[dict], top_n: int = 3) -> list[dict]:
    scored = sorted(
        ({"course_id": str(c["id"]), "course_name": c.get("name", ""), "score": round(_score(materia_name, c.get("name", "")), 3)}
         for c in courses),
        key=lambda x: x["score"], reverse=True,
    )
    return scored[:top_n]


def _cell(ws, row: int, col: int) -> str:
    return str(ws.cell(row=row, column=col).value or "").strip()


def _find_label_value(ws, label_keywords: list[str], max_row: int = 20, max_col: int = 14) -> str:
    """Busca una celda cuyo texto normalizado contenga alguna de las
    palabras clave como PALABRA COMPLETA (ej. 'CI' no debe matchear dentro
    de 'direCCIón') y devuelve el valor no vacío más cercano en un entorno
    alrededor — el layout real es inconsistente: a veces el valor está
    justo abajo, a veces una fila abajo y una columna a la izquierda (ej.
    la etiqueta 'Nombre y Apellido:' en C7 con el nombre real en B8), así
    que se escanea un rectángulo chico en vez de asumir una posición fija."""
    patterns = [re.compile(rf"\b{re.escape(kw)}\b") for kw in label_keywords]
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            val = _cell(ws, r, c)
            if not val:
                continue
            norm = _normalize(val)
            if not any(p.search(norm) for p in patterns):
                continue
            for dr in range(0, 4):
                for dc in range(-2, 4):
                    if dr == 0 and dc <= 0:
                        continue
                    rr, cc = r + dr, c + dc
                    if rr < 1 or cc < 1:
                        continue
                    candidate = _cell(ws, rr, cc)
                    if candidate and not any(p.search(_normalize(candidate)) for p in patterns):
                        return candidate
    return ""


def _find_course_table(ws, max_scan_rows: int = 20, max_scan_cols: int = 20) -> list[str] | None:
    """Busca la fila/columna donde arranca el header 'Curso' y devuelve la
    lista de nombres de materia debajo de esa columna, hasta la primera
    fila vacía. Devuelve None (no lista vacía) si ni siquiera existe la
    tabla — distinción importante: una alumna con la tabla presente pero
    sin filas cargadas todavía es un caso real a mostrar en la vista previa,
    no lo mismo que una hoja que no es de un alumno."""
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        for c in range(1, min(max_scan_cols, ws.max_column) + 1):
            if _normalize(_cell(ws, r, c)) == _CURSO_HEADER:
                materias = []
                row = r + 1
                while row <= ws.max_row:
                    val = _cell(ws, row, c)
                    if not val:
                        break
                    materias.append(val)
                    row += 1
                return materias
    return None


class MateriaMatch(BaseModel):
    materia_original: str
    status: str  # "auto" | "revisar" | "sin_match"
    course_id: str | None = None
    course_name: str | None = None
    score: float | None = None
    candidates: list[dict] = []


class AlumnoMaterias(BaseModel):
    sheet_name: str
    alumno: str
    cedula: str
    materias: list[MateriaMatch]


class MateriasPreviewResponse(BaseModel):
    alumnos: list[AlumnoMaterias]
    skipped_sheets: list[str] = []
    sin_materias: list[dict] = []
    total_alumnos: int = 0
    total_materias: int = 0
    auto_count: int = 0
    review_count: int = 0
    no_match_count: int = 0


@router.post("/preview", response_model=MateriasPreviewResponse, summary="Analizar planilla (una hoja por alumno) y sugerir matches de materias")
async def preview_materias(req: UrlOnlyRequest) -> MateriasPreviewResponse:
    if not req.url or "http" not in req.url:
        raise HTTPException(status_code=400, detail="URL inválida.")

    encoded_url = _encode_share_url(req.url)
    try:
        contents = await graph.get_raw(f"/shares/{encoded_url}/driveItem/content")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo descargar el archivo de OneDrive. {e}")

    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo descargado no es un Excel válido.")

    try:
        courses = await canvas.paginate(
            f"/accounts/{_ACCOUNT}/courses",
            params={"state[]": ["available", "unpublished", "created", "claimed"]},
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo obtener la lista de cursos de Canvas. {e}")

    if not courses:
        raise HTTPException(status_code=400, detail="No se encontraron cursos en la cuenta de Canvas para comparar.")

    alumnos: list[AlumnoMaterias] = []
    skipped: list[str] = []
    sin_materias: list[dict] = []
    auto_count = review_count = no_match_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cedula_raw = _find_label_value(ws, _CEDULA_LABELS)
        cedula = re.sub(r"[.\-\s]", "", cedula_raw)
        alumno = _find_label_value(ws, _NOMBRE_LABELS) or sheet_name.strip()
        materias_raw = _find_course_table(ws)

        if not cedula or materias_raw is None:
            # No se reconoce como hoja de alumno (no tiene cédula, o ni
            # siquiera existe la tabla de Curso/Día/Hora) — ej. la hoja
            # "PRIMER SEMESTRE" con el temario general, no de un alumno.
            skipped.append(sheet_name)
            continue

        if not materias_raw:
            # Es una hoja de alumno real (tiene cédula) pero la tabla de
            # materias está vacía todavía — no se pierde, se informa aparte
            # para que se sepa que falta cargarle materias.
            sin_materias.append({"sheet_name": sheet_name, "alumno": alumno, "cedula": cedula})
            continue

        materia_matches: list[MateriaMatch] = []
        for materia_name in materias_raw:
            candidates = _best_matches(materia_name, courses)
            best = candidates[0] if candidates else None
            if best and best["score"] >= AUTO_THRESHOLD:
                status = "auto"
                auto_count += 1
            elif best and best["score"] >= REVIEW_THRESHOLD:
                status = "revisar"
                review_count += 1
            else:
                status = "sin_match"
                no_match_count += 1

            materia_matches.append(MateriaMatch(
                materia_original=materia_name,
                status=status,
                course_id=best["course_id"] if best and status != "sin_match" else None,
                course_name=best["course_name"] if best and status != "sin_match" else None,
                score=best["score"] if best else None,
                candidates=candidates,
            ))

        alumnos.append(AlumnoMaterias(
            sheet_name=sheet_name, alumno=alumno, cedula=cedula, materias=materia_matches,
        ))

    return MateriasPreviewResponse(
        alumnos=alumnos,
        skipped_sheets=skipped,
        sin_materias=sin_materias,
        total_alumnos=len(alumnos),
        total_materias=sum(len(a.materias) for a in alumnos),
        auto_count=auto_count,
        review_count=review_count,
        no_match_count=no_match_count,
    )


class MateriaEnrollItem(BaseModel):
    cedula: str
    alumno: str
    course_id: str
    materia_original: str


class MateriasEnrollRequest(BaseModel):
    items: list[MateriaEnrollItem]


MAX_ENROLL_ITEMS = 300


@router.post("/enroll", summary="Matricular alumnos en las materias confirmadas")
async def enroll_materias(req: MateriasEnrollRequest, bg_tasks: BackgroundTasks) -> dict:
    """Recibe la lista YA REVISADA por el usuario (después de corregir los
    matches dudosos en la vista previa) y matricula a cada alumno, por
    cédula (SIS ID), en el curso confirmado. Corre en background con
    seguimiento en el Historial de Trabajos, igual que el resto de las
    matriculaciones masivas."""
    items = req.items
    if not items:
        raise HTTPException(status_code=400, detail="No hay matrículas para procesar.")
    if len(items) > MAX_ENROLL_ITEMS:
        raise HTTPException(status_code=400, detail=f"Demasiados registros ({len(items)}). Máximo {MAX_ENROLL_ITEMS} por ejecución.")

    job_id = await jobs.create_job(job_type="materias_masivo", operation="import", username="admin")
    bg_tasks.add_task(_process_materias_enroll_bg, job_id, items)
    return {"status": "success", "job_id": job_id, "message": "Matriculación por materias iniciada en segundo plano."}


async def _process_materias_enroll_bg(job_id: int, items: list[MateriaEnrollItem]):
    await jobs.start_job(job_id)

    success_count = 0
    error_count = 0
    results: list[dict] = []
    user_id_cache: dict[str, str | None] = {}

    async def get_canvas_user_id(cedula: str) -> str | None:
        if cedula in user_id_cache:
            return user_id_cache[cedula]
        try:
            u = await canvas.get(f"/accounts/{_ACCOUNT}/users/sis_user_id:{cedula}")
            uid = str(u["id"]) if u and u.get("id") else None
        except Exception:
            uid = None
        user_id_cache[cedula] = uid
        return uid

    async def process_item(item: MateriaEnrollItem):
        nonlocal success_count, error_count
        uid = await get_canvas_user_id(item.cedula)
        if not uid:
            error_count += 1
            results.append({
                "alumno": item.alumno, "materia": item.materia_original,
                "status": f"❌ Alumno no encontrado en Canvas (cédula {item.cedula})",
            })
            return
        try:
            await canvas.post(f"/courses/{item.course_id}/enrollments", {
                "enrollment": {
                    "user_id": uid, "type": "StudentEnrollment",
                    "enrollment_state": "active", "notify": False,
                },
            })
            success_count += 1
            results.append({"alumno": item.alumno, "materia": item.materia_original, "status": "✅ OK"})
        except Exception as e:
            msg = str(e)
            if "already" in msg.lower() or "enrolled" in msg.lower():
                success_count += 1
                results.append({"alumno": item.alumno, "materia": item.materia_original, "status": "✅ Ya estaba matriculado"})
            else:
                error_count += 1
                results.append({"alumno": item.alumno, "materia": item.materia_original, "status": f"❌ {msg}"})

    batch_size = 5
    for i in range(0, len(items), batch_size):
        chunk = items[i:i + batch_size]
        await asyncio.gather(*(process_item(it) for it in chunk))
        await jobs.update_job_progress(
            job_id, success_count, error_count,
            data_json=json.dumps({"total_to_process": len(items), "processed": success_count + error_count, "results": results}),
        )

    if error_count and success_count == 0:
        await jobs.fail_job(job_id, f"Todas las matrículas fallaron ({error_count}).")
    else:
        await jobs.complete_job(job_id, success_count, error_count)
