"""Onboarding de nuevos alumnos y docentes.

Flujo principal:
  1. Generar credenciales institucionales a partir del nombre + cédula.
  2. Crear cuenta en Canvas LMS (si platform incluye 'canvas').
  3. Crear cuenta en Azure AD / Microsoft Teams (si platform incluye 'teams').
  4. Enviar email de bienvenida con las credenciales (si send_email=True).

Endpoints públicos:
  POST /ingreso/preview             – Previsualizar credenciales sin crear nada.
  POST /ingreso/create              – Crear usuario individual.
  POST /ingreso/bulk                – Crear múltiples usuarios en paralelo.
  POST /ingreso/check-account       – Verificar si un usuario existe en Canvas/Teams.
  POST /ingreso/bulk-check          – Verificar múltiples usuarios.
  GET  /ingreso/template/crear      – Descargar plantilla Excel.
  POST /ingreso/bulk-file           – Crear usuarios desde archivo Excel.
"""
import asyncio
import re
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, File, UploadFile
from pydantic import BaseModel, field_validator

from app.core.config import settings
from app.models.canvas import BulkResult
from app.services import canvas_client as canvas
from app.services import teams_client as graph
from app.services import user_service

router = APIRouter(prefix="/ingreso", tags=["Nuevo Ingreso"])
_ACCOUNT = settings.canvas_account_id

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
_VALID_PLATFORMS = {"canvas", "teams", "both"}
_VALID_PROGRAMS  = {"grado", "mba", "diplomado"}
_VALID_ROLES     = {"student", "teacher"}


# ── Modelos ────────────────────────────────────────────────────────────────────

class StudentIn(BaseModel):
    """Datos necesarios para crear un nuevo usuario en Canvas y/o Teams."""

    full_name: str
    cedula: str
    personal_email: str
    role: str = "student"
    platform: str = "both"
    program_type: str = "grado"
    program_name: str = ""
    send_email: bool = True
    cc: list[str] = []
    courses: list[str] = []

    @field_validator("full_name")
    @classmethod
    def v_full_name(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("El nombre completo es requerido")
        if len(v) < 3:
            raise ValueError("El nombre debe tener al menos 3 caracteres")
        return v

    @field_validator("cedula")
    @classmethod
    def v_cedula(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("La cédula es requerida")
        cleaned = v.replace("-", "").replace(".", "")
        if not cleaned.isdigit():
            raise ValueError("La cédula debe contener solo números")
        return v

    @field_validator("personal_email")
    @classmethod
    def v_email(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("El email personal es requerido")
        if not _EMAIL_RE.match(v):
            raise ValueError(f"El email '{v}' no tiene formato válido")
        return v

    @field_validator("role")
    @classmethod
    def v_role(cls, v: str) -> str:
        v = v.strip().lower()
        if v not in _VALID_ROLES:
            raise ValueError(f"Rol inválido '{v}'. Use: {_VALID_ROLES}")
        return v

    @field_validator("platform")
    @classmethod
    def v_platform(cls, v: str) -> str:
        v = v.strip().lower()
        if v not in _VALID_PLATFORMS:
            raise ValueError(f"Plataforma inválida '{v}'. Use: {_VALID_PLATFORMS}")
        return v

    @field_validator("program_type")
    @classmethod
    def v_program_type(cls, v: str) -> str:
        v = v.strip().lower()
        if v not in _VALID_PROGRAMS:
            raise ValueError(f"Tipo de programa inválido '{v}'. Use: {_VALID_PROGRAMS}")
        return v


class BulkStudentsIn(BaseModel):
    """Lista de usuarios para creación conjunta."""
    students: list[StudentIn]


class ResendCredentialsIn(BaseModel):
    """Datos para reenviar credenciales a un usuario ya existente."""

    cedula: str
    personal_email: str
    full_name: str
    platform: str = "both"
    program_type: str = "grado"
    program_name: str = ""
    cc: list[str] = []

    @field_validator("cedula")
    @classmethod
    def v_cedula(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("La cédula es requerida")
        return v

    @field_validator("personal_email")
    @classmethod
    def v_email(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("El email personal es requerido")
        if not _EMAIL_RE.match(v):
            raise ValueError(f"El email '{v}' no tiene formato válido")
        return v

    @field_validator("full_name")
    @classmethod
    def v_full_name(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("El nombre completo es requerido")
        return v


class BulkResendIn(BaseModel):
    """Lista de usuarios para reenvío conjunto."""
    students: list[ResendCredentialsIn]


class AccountCheckIn(BaseModel):
    """Datos para verificar existencia de un usuario en Canvas/Teams."""

    cedula: str
    full_name: str = ""
    platform: str = "both"

    @field_validator("cedula")
    @classmethod
    def v_cedula(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("La cédula es requerida para la verificación")
        return v


class BulkAccountCheckIn(BaseModel):
    """Lista de usuarios para verificación conjunta."""
    students: list[AccountCheckIn]


class CredentialPreview(BaseModel):
    """Respuesta de previsualización de credenciales generadas."""

    full_name: str
    cedula: str
    personal_email: str
    role: str
    login_id: str
    institutional_email: str
    password: str


# ── Helpers internos ───────────────────────────────────────────────────────────

def _err(exc: Exception) -> str:
    """Extrae el mensaje de error más útil de cualquier excepción."""
    return exc.detail if isinstance(exc, HTTPException) else str(exc)


def _resolve_login(creds: dict, role: str) -> tuple[str, str]:
    """Devuelve (canvas_login_id, canvas_sis_user_id).

    Tanto estudiantes como docentes usan el email institucional como login_id.
    El SIS user ID es la cédula para estudiantes, y el email para docentes.
    """
    if role == "teacher":
        return creds["email"], creds["email"]
    return creds["email"], creds["cedula"]



async def _check_account(body: AccountCheckIn) -> dict[str, Any]:
    """Verifica si un usuario existe en Canvas y/o Teams.

    Lógica de búsqueda:
    - Canvas: por cédula (SIS ID), fallback por email institucional generado.
    - Teams: por userPrincipalName (email institucional generado).

    Si no se provee full_name, Teams no puede buscarse (el UPN se deriva del nombre).
    """
    result: dict[str, Any] = {
        "cedula": body.cedula,
        "full_name": body.full_name,
    }

    if body.full_name.strip():
        creds, status = await user_service.generate_unique_credentials(body.full_name, body.cedula, platform=body.platform)
        login_id = creds["email"]
        upn = creds["email"]
        result["generated_email"] = creds["email"]
    else:
        login_id = body.cedula
        upn = ""

    if body.platform in ("canvas", "both"):
        try:
            exists, info = await user_service._canvas_user_exists(body.cedula, login_id)
            result["canvas"] = {"exists": exists, **(info if exists else {})}
        except Exception as exc:
            result["canvas"] = {"exists": None, "error": _err(exc)}

    if body.platform in ("teams", "both"):
        if not upn:
            result["teams"] = {
                "exists": None,
                "error": "Se requiere nombre completo para buscar en Teams",
            }
        else:
            try:
                exists, info = await _teams_user_exists(upn)
                result["teams"] = {"exists": exists, **(info if exists else {})}
            except Exception as exc:
                result["teams"] = {"exists": None, "error": _err(exc)}

    return result


async def _generate_unique_credentials(full_name: str, cedula: str, platform: str) -> dict:
    creds = generate_credentials(full_name, cedula, settings.institutional_domain)
    
    # 1. Verificar si ya existe por cédula (no cambiar correo si es el mismo usuario)
    try:
        if platform in ("canvas", "both"):
            exists, info = await _canvas_user_exists(cedula, creds["email"])
            if exists and info.get("found_by") == "cedula":
                return creds
    except Exception:
        pass
async def _create_student(student: StudentIn) -> dict[str, Any]:
    """Crea un usuario en Canvas y/o Teams según la plataforma indicada.

    Flujo:
    1. Genera credenciales institucionales únicas (manejando colisiones).
    2. Pre-verifica existencia antes de intentar crear (evita duplicados).
    3. Crea en Canvas si platform in ('canvas', 'both').
    4. Crea en Teams/Azure AD si platform in ('teams', 'both').
    5. Envía email de bienvenida si send_email=True.
    6. Matricula en los cursos de Canvas si se proveyeron en `courses`.

    Returns:
        Diccionario con estado de cada plataforma y las credenciales generadas.
    """
    creds, _ = await user_service.generate_unique_credentials(student.full_name, student.cedula, platform=student.platform)
    login_id, sis_user_id = _resolve_login(creds, student.role)
    results: dict[str, Any] = {
        "student": student.full_name,
        "role": student.role,
        "credentials": {**creds, "login_id": login_id},
    }

    # ── Canvas ────────────────────────────────────────────────
    if student.platform in ("canvas", "both"):
        try:
            exists, info = await user_service._canvas_user_exists(sis_user_id, login_id)
            if exists:
                results["canvas"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Canvas (registrado como: {info.get('name', '?')})",
                    "existing": info,
                }
            else:
                canvas_user = await canvas.post(
                    f"/accounts/{_ACCOUNT}/users",
                    {
                        "user": {
                            "name": creds["full_name"],
                            "short_name": creds["full_name"],
                        },
                        "pseudonym": {
                            "unique_id": login_id,
                            "sis_user_id": sis_user_id,
                            "password": creds["password"],
                            "send_confirmation": False,
                        },
                        "communication_channel": {
                            "type": "email",
                            "address": creds["email"],
                            "skip_confirmation": True,
                        },
                    },
                )
                results["canvas"] = {"status": "ok", "id": canvas_user.get("id")}
        except Exception as exc:
            error_str = _err(exc)
            if any(k in error_str.lower() for k in ("unique_id", "taken", "already")):
                results["canvas"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Canvas: {error_str}",
                }
            else:
                results["canvas"] = {"status": "error", "error": error_str}

    # ── Teams / Azure AD ──────────────────────────────────────
    if student.platform in ("teams", "both"):
        parts = student.full_name.strip().split()
        sku = settings.azure_sku_teachers if student.role == "teacher" else settings.azure_sku_students
        try:
            exists, info = await user_service._teams_user_exists(creds["email"])
            if exists:
                results["teams"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Teams (registrado como: {info.get('name', '?')})",
                    "existing": info,
                }
            else:
                az_user = await graph.post(
                    "/users",
                    {
                        "displayName": creds["full_name"],
                        "givenName": parts[0],
                        "surname": " ".join(parts[1:]) if len(parts) > 1 else "",
                        "userPrincipalName": creds["email"],
                        "mailNickname": creds["login_id"].replace(".", "_"),
                        "usageLocation": settings.usage_location,
                        "accountEnabled": True,
                        "passwordProfile": {
                            "forceChangePasswordNextSignIn": True,
                            "password": creds["password"],
                        },
                    },
                )
                await graph.assign_license(az_user["id"], sku)
                results["teams"] = {
                    "status": "ok",
                    "id": az_user.get("id"),
                    "license": sku,
                }
        except Exception as exc:
            error_str = _err(exc)
            if any(k in error_str for k in ("ObjectConflict", "conflictingObjects")) or \
               "already exists" in error_str.lower():
                results["teams"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Teams: {error_str}",
                }
            else:
                results["teams"] = {"status": "error", "error": error_str}



    # ── Canvas Enrollments ──────────────────────────────────────
    if student.courses and results.get("canvas", {}).get("status") in ("ok", "exists"):
        canvas_id = results["canvas"].get("id") or results["canvas"].get("existing", {}).get("canvas_id")
        if canvas_id:
            results["canvas_enrollments"] = []
            for course_id in student.courses:
                try:
                    await canvas.post(
                        f"/courses/{course_id}/enrollments",
                        {
                            "enrollment": {
                                "user_id": canvas_id,
                                "type": "StudentEnrollment" if student.role == "student" else "TeacherEnrollment",
                                "enrollment_state": "invited",
                                "notify": True
                            }
                        }
                    )
                    results["canvas_enrollments"].append({"course_id": course_id, "status": "success"})
                except Exception as exc:
                    results["canvas_enrollments"].append({"course_id": course_id, "status": "error", "error": _err(exc)})

    return results




# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("/courses", summary="Obtener cursos activos de Canvas")
async def get_courses(search: str = ""):
    """Busca cursos en Canvas para la selección de materias."""
    try:
        params = {"published": True, "per_page": 50}
        if search:
            params["search_term"] = search
        courses = await canvas.get(f"/accounts/{_ACCOUNT}/courses", params)
        if isinstance(courses, list):
            return [{"id": str(c["id"]), "name": c["name"]} for c in courses if "name" in c]
        return []
    except Exception:
        return []


@router.post("/preview", response_model=CredentialPreview, summary="Previsualizar credenciales")
async def preview_credentials(body: StudentIn):
    """Genera y muestra las credenciales que se asignarían al usuario, sin crear nada."""
    creds, _ = await user_service.generate_unique_credentials(body.full_name, body.cedula, body.platform)
    login_id, _ = _resolve_login(creds, body.role)
    return CredentialPreview(
        full_name=body.full_name,
        cedula=body.cedula,
        personal_email=body.personal_email,
        role=body.role,
        login_id=login_id,
        institutional_email=creds["email"],
        password=creds["password"],
    )


@router.post("/create", summary="Crear credenciales para un alumno o docente")
async def create_student(body: StudentIn):
    """Crea un usuario en Canvas y/o Teams y envía el correo de bienvenida."""
    return await _create_student(body)


@router.post("/bulk", summary="Crear credenciales conjuntas")
async def create_students_bulk(body: BulkStudentsIn) -> BulkResult:
    """Crea múltiples usuarios en paralelo. Errores individuales no detienen el proceso."""
    result = BulkResult()

    async def _run(student: StudentIn):
        try:
            data = await _create_student(student)
            result.succeeded.append(data)
        except Exception as exc:
            result.failed.append({"student": student.full_name, "error": str(exc)})

    await asyncio.gather(*[_run(s) for s in body.students])
    return result


@router.post("/check-account", summary="Verificar si un usuario existe en Canvas y/o Teams")
async def check_account(body: AccountCheckIn) -> dict[str, Any]:
    """Verifica si un usuario ya existe antes de intentar crearlo."""
    return await _check_account(body)


@router.post("/bulk-check", summary="Verificar cuentas masivamente")
async def check_accounts_bulk(body: BulkAccountCheckIn) -> dict[str, Any]:
    """Verifica la existencia de múltiples usuarios en paralelo y devuelve estadísticas."""
    results = await asyncio.gather(
        *[_check_account(s) for s in body.students],
        return_exceptions=True,
    )
    output = []
    for body_item, res in zip(body.students, results):
        if isinstance(res, Exception):
            output.append({
                "cedula": body_item.cedula,
                "full_name": body_item.full_name,
                "error": str(res),
            })
        else:
            output.append(res)

    return {
        "total": len(output),
        "found_canvas": sum(
            1 for r in output
            if isinstance(r, dict) and r.get("canvas", {}).get("exists") is True
        ),
        "found_teams": sum(
            1 for r in output
            if isinstance(r, dict) and r.get("teams", {}).get("exists") is True
        ),
        "results": output,
    }


@router.get("/template/crear", summary="Descargar plantilla Excel para creación conjunta")
async def template_crear():
    """Genera y descarga una plantilla Excel con el formato requerido para carga conjunta."""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = [
        "Nombre Completo", "Cedula", "Email Personal",
        "Rol", "Plataforma", "Programa", "Nombre del Programa",
    ]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font  = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill  = openpyxl.styles.PatternFill(
                start_color="5A67D8", end_color="5A67D8", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    ws.append(["Karen Gonzalez", "6868066", "karen@gmail.com", "student", "both", "grado", ""])
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_crear_usuarios.xlsx"},
    )


@router.post("/bulk-file", summary="Crear usuarios desde archivo Excel")
async def bulk_file_create(file: UploadFile = File(...)) -> BulkResult:
    """Lee un archivo Excel (.xlsx/.xls) y crea usuarios masivamente.

    Columnas esperadas (en orden):
    A: Nombre Completo | B: Cédula | C: Email Personal
    D: Rol             | E: Plataforma | F: Programa | G: Nombre del Programa
    """
    import io
    import openpyxl

    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {exc}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="El archivo no contiene datos. Agregue al menos una fila después del encabezado.",
        )

    result = BulkResult()

    async def _create_from_row(row: tuple):
        name = str(row[0]).strip() if row[0] else ""
        cedula = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not name or not cedula:
            result.failed.append({"student": name or "?", "error": "Nombre y cédula son requeridos"})
            return
        try:
            student_data = StudentIn(
                full_name=name,
                cedula=cedula,
                personal_email=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                role=str(row[3]).strip().lower() if len(row) > 3 and row[3] else "student",
                platform=str(row[4]).strip().lower() if len(row) > 4 and row[4] else "both",
                program_type=str(row[5]).strip().lower() if len(row) > 5 and row[5] else "grado",
                program_name=str(row[6]).strip() if len(row) > 6 and row[6] else "",
                send_email=True,
                cc=[],
            )
            data = await _create_student(student_data)
            result.succeeded.append(data)
        except Exception as exc:
            result.failed.append({"student": name, "error": _err(exc)[:300]})

    await asyncio.gather(*[_create_from_row(row) for row in rows])
    return result


@router.post("/bulk-preview", summary="Previsualizar creación conjunta")
async def preview_students_bulk(body: BulkStudentsIn) -> dict:
    """Simula la creación de múltiples usuarios y retorna qué credenciales y estados tendrían."""
    sample = []
    
    async def _preview(student: StudentIn):
        creds, status = await user_service.generate_unique_credentials(
            student.full_name, student.cedula, platform=student.platform
        )
        login_id, sis_user_id = _resolve_login(creds, student.role)
        
        status_canvas = "-"
        status_teams = "-"
        
        if student.platform in ("canvas", "both"):
            exists, _ = await user_service._canvas_user_exists(sis_user_id, login_id)
            status_canvas = "exists" if exists else "new"
            
        if student.platform in ("teams", "both"):
            exists, _ = await user_service._teams_user_exists(creds["email"])
            status_teams = "exists" if exists else "new"
            
        sample.append({
            "nombre": student.full_name,
            "cedula": student.cedula,
            "correo_personal": student.personal_email,
            "login_id": login_id,
            "correo_institucional": creds["email"],
            "password": creds["password"],
            "plataforma": student.platform,
            "status_canvas": status_canvas,
            "status_teams": status_teams
        })

    import asyncio
    await asyncio.gather(*[_preview(s) for s in body.students])
    
    return {
        "total_rows": len(body.students),
        "valid_rows": len(body.students),
        "sample": sample
    }
